from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
from ttkbootstrap.constants import *
import ttkbootstrap as tb

import cv2
import openpyxl
from pyzbar.pyzbar import decode



# Creating a window
root = tb.Window(themename='litera')
root.title("Login Page")
root.geometry("450x680")


# Load and Set Window Icon
logo_img = Image.open("logo.png")
logo_img = ImageTk.PhotoImage(logo_img)
root.iconphoto(False, logo_img)  # Set the window icon


# Load and resize the image
image = Image.open('gitam-Logo-new.png')
image = image.resize((250, 110), Image.LANCZOS)
photo = ImageTk.PhotoImage(image)


# Logo Label (Placed above the frame)
logo_label = Label(root, image=photo, bg=root["bg"])
logo_label.pack(pady=20)


# Create a frame for the login elements
login_frame = tb.Frame(root, borderwidth=10, relief='ridge')
login_frame.pack(expand=True, padx=10, pady=10)


# Login Page elements
label1 = tb.Label(login_frame, text="Gitam Mail", font=("Arial", 14))
label1.pack(padx=10, pady=10)
entry1 = tb.Entry(login_frame, font=("Arial", 14), width=20)
entry1.pack(padx=10, pady=10)

label2 = tb.Label(login_frame, text="Password", font=("Arial", 14))
label2.pack(padx=10, pady=10)
entry2 = tb.Entry(login_frame, show="*", font=("Arial", 14), width=20)
entry2.pack(padx=10, pady=10)



# to get user data from the excel workbook
def get_users_data():
    try:
        wb = openpyxl.load_workbook("users.xlsx")
        sheet = wb.active
        return list(sheet.iter_rows(values_only=True))  # Get all data
    except FileNotFoundError:
        return []

# showing the login screen
def show_login_screen():
    # Clear everything before reloading login UI
    for widget in root.winfo_children():
        widget.destroy()


logged_in_email = None  # Global variable to store the email

# function to login as a user 
def login():
    global logged_in_email
    name = entry1.get().strip()
    password = entry2.get().strip()

    try:
        wb = openpyxl.load_workbook("users.xlsx")
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]  # Read first row (header)
        
        try:
            name_col = headers.index("Username")
            pass_col = headers.index("Password")
            balance_col = headers.index("Balance")
        except ValueError:
            messagebox.showerror("Error", "Required columns not found in the Excel file.")
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row[name_col] == name and row[pass_col] == password:
                try:
                    balance = float(row[balance_col])  # Convert balance to float
                except (ValueError, TypeError):
                    messagebox.showerror("Error", f"Invalid balance value: {row[balance_col]}")
                    return

                logged_in_email = name  # Store the logged-in email globally
                messagebox.showinfo("Login Status", "Login Successful")
                dash(balance)  # Open dashboard with balance
                return

        messagebox.showerror("Login Status", "Invalid Credentials")

    except FileNotFoundError:
        messagebox.showerror("Error", "User database not found!")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {e}")


# function to register a new user
def register_user():
    hex_value = entry5.get().strip()
    new_gitam_mail = entry3.get().strip()
    new_password = entry4.get()
    confirm_password = entry6.get()

    # Ensure all fields are filled
    if not hex_value or not new_gitam_mail or not new_password or not confirm_password:
        messagebox.showerror("Registration Status", "All fields are required!")
        return

    if new_password != confirm_password:
        messagebox.showerror("Registration Status", "Passwords do not match!")
        return

    users_data = get_users_data()

    for row in users_data:
        if len(row) < 3:
            continue
        existing_email, _, existing_hex = row[:3]
        if new_gitam_mail == existing_email:
            messagebox.showerror("Registration Status", "GITAM Mail already exists!")
            return
        if hex_value == existing_hex:
            messagebox.showerror("Registration Status", "This RFID is already registered!")
            return

    try:
        wb = openpyxl.load_workbook("users.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Gitam Mail", "Password", "Hex Value", "Service", "Balance"])  # Add 'Service' column
    
    # Append new user with default values:
    sheet.append([new_gitam_mail, new_password, hex_value, True, 0])
    wb.save("users.xlsx")
    
    messagebox.showinfo("Registration Status", "User registered successfully")
    #show_login_screen()  # Navigate back to login
    root.destroy()


def show_register_screen():
    login_frame.pack_forget()
    register_frame.pack(expand=True)




style = tb.Style()
style.configure("Custom.TButton",
                background="#006B64", foreground="white", font=("Arial", 12), padding=10, borderwidth=0, focuscolor="")

button = tb.Button(login_frame, text='Login', command=login, style="Custom.TButton")
button.pack(pady=25)

register_button = tb.Button(login_frame, text="Register", command=show_register_screen, style="Custom.TButton")
register_button.pack(pady=20)

register_frame = tb.Frame(root)
label5 = tb.Label(register_frame, text="RFID UID", font=("Arial", 14))
label5.pack(padx=10, pady=10)
entry5 = tb.Entry(register_frame, font=("Arial", 14), width=20)
entry5.pack(padx=10, pady=10)

label3 = tb.Label(register_frame, text="Gitam Mail", font=("Arial", 14))
label3.pack(padx=10, pady=10)
entry3 = tb.Entry(register_frame, font=("Arial", 14), width=20)
entry3.pack(padx=10, pady=10)

label4 = tb.Label(register_frame, text="Password", font=("Arial", 14))
label4.pack(padx=10, pady=10)
entry4 = tb.Entry(register_frame, show="*", font=("Arial", 14), width=20)
entry4.pack(padx=10, pady=10)

label6 = tb.Label(register_frame, text="Confirm Password", font=("Arial", 14))
label6.pack(padx=10, pady=10)
entry6 = tb.Entry(register_frame, show="*", font=("Arial", 14), width=20)
entry6.pack(padx=10, pady=10)

register_submit_button = tb.Button(register_frame, text="Register", command=register_user, style="Custom.TButton")
register_submit_button.pack(pady=20)



# using computer vision module to scan the qr and get the data in it
def scan_qr_code():
    global logged_in_email
    cap = cv2.VideoCapture(0)
    detector = cv2.QRCodeDetector()

    while True:
        _, frame = cap.read()
        qr_data, bbox, _ = detector.detectAndDecode(frame)

        if qr_data:
            cap.release()
            cv2.destroyAllWindows()
            
            if not logged_in_email:
                messagebox.showerror("Error", "User email not found. Please log in again.")
                return
            
            process_qr_payment(qr_data, logged_in_email)  # Pass the stored email
            return

        cv2.imshow("Scan QR Code", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


# from the data from the previous function, payment is processed
def process_qr_payment(qr_data, user_email):
    try:
        amount = float(qr_data.strip())  # Extract amount from QR
        wb = openpyxl.load_workbook("users.xlsx")
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        try:
            email_col = headers.index("Username")
            balance_col = headers.index("Balance")  # Dynamically find the balance column
        except ValueError:
            messagebox.showerror("Error", "Required columns not found in the Excel file.")
            return

        # comparing the input valus to the workbook values
        for row in sheet.iter_rows(min_row=2, values_only=False):  # Skip header row
            if row[email_col].value == user_email:
                balance_cell = row[balance_col]
                current_balance = float(balance_cell.value)

                if current_balance >= amount:
                    balance_cell.value = current_balance - amount
                    wb.save("users.xlsx")
                    messagebox.showinfo("Payment Successful", f"₹{amount} deducted from your balance.")
                    return
                else:
                    messagebox.showerror("Payment Failed", "Insufficient Balance!")
                    return

        messagebox.showerror("Error", "User not found!")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


def on_pay_click():
    scan_qr_code()



# function to display the dashboard of a particulat user
def dash(bal):
    # Clear the entire window
    # to load the dashboard in the same login window
    for widget in root.winfo_children():
        widget.destroy()

    # Set new background color
    root.configure(bg="#F5E1BE")

    # Define a custom style for the top frame
    style = tb.Style()
    style.configure("Custom.TFrame", background="#006B64")
    style.configure("Custom.TLabel", foreground="white", background="#006B64", font=("Helvica", 22, "bold"))
    style.configure("Custom.TButton", font=("Arial", 14, "bold"), padding=(15, 8), width=12)

    # Top section with specified background color
    top_frame = tb.Frame(root, style="Custom.TFrame", height=75)
    top_frame.pack(fill='x')
    top_frame.pack_propagate(False)

    # Logout button (top left)
    red_button = tb.Button(top_frame, text="⏻", bootstyle="danger", command=root.destroy)
    red_button.place(x=20, y=20)

    # Title label
    title_label = tb.Label(top_frame, text="G - Wallet", style="Custom.TLabel")
    title_label.pack(expand=True)

    # White rectangular frame for balance
    balance_frame = tb.Frame(root, bootstyle="light", padding=10, borderwidth=3, relief="solid")
    balance_frame.pack(pady=15, padx=50, fill='x')

    balance_label = tb.Label(balance_frame, text=f"Balance: ₹{bal}", font=("Arial", 18))
    balance_label.pack(pady=10)


    # settings frame
    settings_frame = tb.Frame(root, width=410, height=580, bootstyle="light")
    settings_frame.pack(pady=30)
    settings_frame.configure(borderwidth=2, relief="solid") 
    settings_frame.pack_propagate(False)

    # Pay Button
    pay_button = tb.Button(settings_frame, text="Pay Using QR", command=on_pay_click, style="Custom.TButton")
    pay_button.pack(pady=25)


    # checkboxes to add functionalities to the dashboard
    style.configure('Custom.TCheckbutton', font=('Arial', 14), padding=(20, 5), width=20, anchor='c')

    var1 = IntVar(value=1)
    checkbox1 = tb.Checkbutton(settings_frame, text="Enable RFID Payment",style="Custom.TCheckbutton", variable=var1, onvalue=1, offvalue=0)
    checkbox1.pack(pady=10)
    
    var2 = IntVar(value=1)
    checkbox2 = tb.Checkbutton(settings_frame, text="Enable OTP Status",style="Custom.TCheckbutton", variable=var2, onvalue=1, offvalue=0)
    checkbox2.pack(pady=10)


    # powered label
    powered_by = Label(settings_frame, text="Powered by CATS", font=("Arial", 10), fg='white', bg='#F5E1BE')
    powered_by.pack(side='bottom', pady=10)


root.mainloop()
