import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
import ttkbootstrap as tb
import openpyxl


# Function to check and update Excel file dynamically
def proceed_action():
    email = email_entry.get().strip()
    amount = amount_entry.get().strip()

    # getting both mail, amount input from the G-Pay window
    if not email or not amount:
        messagebox.showerror("Error", "All fields are required!")
        return

    # converting the amount into float 
    try:
        amount = float(amount)
    except ValueError:
        messagebox.showerror("Error", "Amount must be a valid number!")
        return

    file_path = "users.xlsx"  # Ensure the file is in the same directory or update the path

    # loading all the elements of the excel sheet into an active workbook
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Assuming data is in the first sheet

        # Find column indices dynamically
        headers = {cell.value: idx for idx, cell in enumerate(sheet[1]) if cell.value}

        # display an error when any attributes are missing
        if "Username" not in headers or "Balance" not in headers:
            messagebox.showerror("Error", "Required columns (Email, Balance) not found!")
            return

        # getting all the usernames and the balances 
        email_col = headers["Username"]
        balance_col = headers["Balance"]

        user_found = False

        # checking for the required username and balance to make changes
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[email_col].value == email:
                try:
                    current_balance = float(row[balance_col].value)
                except ValueError:
                    messagebox.showerror("Error", "Invalid balance format in the file!")
                    return
                
                row[balance_col].value = current_balance + amount  # Update balance
                user_found = True
                break

        # when the user is found and the amount was added successfully 
        if user_found:
            workbook.save(file_path)
            messagebox.showinfo("Success", "Amount added successfully!")
        else:
            messagebox.showerror("Error", "User not found!")

    # handeling some exceptions 
    except FileNotFoundError:
        messagebox.showerror("Error", "users.xlsx file not found!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")



# Main window
root = tb.Window()
root.title("Gitam Payment Portal")
root.geometry("500x280")
root.resizable(False, False)

# Load and Set Window Icon(top left icon)
logo_img = Image.open("logo.png")
logo_img = ImageTk.PhotoImage(logo_img)
root.iconphoto(False, logo_img)


# Canvas for background rectangle(for the heading)
canvas = tk.Canvas(root, width=500, height=80, bg="white", highlightthickness=0)
canvas.place(x=0, y=0)
canvas.create_rectangle(0, 0, 600, 200, fill="#006B64", outline="")


# Heading label
style = tb.Style()
style.configure("Custom.TLabel", foreground="white", background="#006B64", font=("Helvica", 16, "bold"))
style.configure("Custom.TButton", foreground="white", background="#006B64",padding=10, borderwidth=0, focuscolor="")
heading = tb.Label(root, text="Gitam Payment Portal", style = "Custom.TLabel")
heading.place(x=120, y=25)


# Gitam mail label & entry
email_label = tb.Label(root, text="Gitam Mail:", font=("Arial", 13))
email_label.place(x=40, y=100)
email_entry = tb.Entry(root, width=30)
email_entry.place(x=200, y=100)

# Amount label & entry
amount_label = tb.Label(root, text="Amount:", font=("Arial", 13))
amount_label.place(x=40, y=140)
amount_entry = tb.Entry(root, width=30)
amount_entry.place(x=200, y=140)

# Proceed button
proceed_button = tb.Button(root, text="Proceed", style="Custom.TButton", command=proceed_action)
proceed_button.place(x=280, y=200)

root.mainloop()
