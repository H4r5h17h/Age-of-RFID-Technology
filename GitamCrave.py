import tkinter as tk
import ttkbootstrap as tb
from tkinter import messagebox
from ttkbootstrap.widgets import Entry, Button, Frame, Label
from tkinter import Toplevel, Label, Button, Frame, StringVar, Canvas, Scrollbar, Entry
from PIL import Image, ImageTk
import openpyxl
import threading
import subprocess
import qrcode
import serial
import time
import os




# Set up serial communication with Arduino
PORT = 'COM3' 
BAUD_RATE = 9600


# function to get the input from the arduino 
def read_rfid(timeout=2):
    try:
        # Attempt to open serial connection
        arduino = serial.Serial(PORT, BAUD_RATE, timeout=1)
        time.sleep(1)  # Allow connection to stabilize
        print("Waiting for RFID scan...")

        start_time = time.time()
        while time.time() - start_time < timeout:
            if arduino.in_waiting > 0:
                hex_value = arduino.readline().decode("utf-8", errors="ignore").strip()
                print(f"✅ RFID Scanned: {hex_value}")
                arduino.close()
                return hex_value  # Return the scanned value

        print("⚠️ No RFID card detected within the timeout period.")
        return None  # RFID module is connected, but no card was scanned

    except serial.SerialException:
        print("❌ No RFID Module Found!")
        return "MODULE_NOT_FOUND"  # RFID module is not connected

    except Exception as e:
        print(f"⚠️ Error reading RFID: {e}")
        return "MODULE_NOT_FOUND"  # Treat other errors as module not found

    finally:
        if 'arduino' in locals() and arduino.is_open:
            arduino.close()



# giving return values to the output from the previous function
def rfid():
    rfid_value = read_rfid()  # Get RFID value

    if rfid_value == "MODULE_NOT_FOUND":
        return "MODULE_NOT_FOUND"  # Only return this when module is missing

    if rfid_value is None or rfid_value.strip() == "":
        return None  # RFID module is there, but no card detected

    return rfid_value.strip()  # Return the actual scanned RFID value




""" this is a sample super market where we can add items to our cart,
    get the bill summary at the end, and finally proceed to pay,
    after clicking this button we have to tap our rfid card to deduct the money"""
def gitam_mart():

    # basic GUI window 
    root = tb.Window(themename="default")
    root.geometry("1300x700")
    root.title("Gitam Crave")

    # adding gitam logo at the top of the window
    logo_img = Image.open("logo.png")
    logo_img = ImageTk.PhotoImage(logo_img)
    root.iconphoto(False, logo_img)

    # the main heading of our window
    header = Frame(root, bg="#006B64", height=60)
    header.pack(fill='x')
    header_content = Frame(header, bg="#006B64")
    header_content.pack(side="top", pady=10)

    # adding a logo beside the heading
    logo_small = Image.open("logo_white.png").resize((75, 70), Image.LANCZOS) # LANCZOS to get HD image
    logo_small = ImageTk.PhotoImage(logo_small)
    Label(header_content, image=logo_small, bg="#006B64").pack(side="left", padx=5)
    Label(header_content, text="Gitam Crave", font=("Arial", 24, "bold"), fg="white", bg="#006B64").pack(side="left", padx=5)

    # our cart to store all the items that we select 
    cart = {}

    # code to update the cart whenever the add button is pressed
    def update_cart(item, price, qty_var):
        qty = int(qty_var.get())
        if qty > 0:
            cart[item] = (price, qty)
        elif item in cart:
            del cart[item]


    # function to show the bill summary for the items in the cart
    def show_bill():
        global total_price
        bill_win = Toplevel()
        bill_win.geometry("460x550")
        bill_win.title("Bill Summary")
        bill_win.configure(bg="white")  # Background remains white
        bill_win.iconphoto(False, logo_img)
    
        # Header with Green Background (Spanning Full Width)
        header = Frame(bill_win, bg="#006B64")
        header.pack(fill="x")  # Fills the entire x-axis
        Label(header, text="Bill Summary", font=("Arial", 16, "bold"), fg="white", bg="#006B64", pady=10).pack()
        
        # Bill Frame with Border
        bill_frame = Frame(bill_win, borderwidth=2, relief="solid", bg="white")
        bill_frame.pack(padx=10, pady=5, fill="x")
        
        # Column Headers
        Label(bill_frame, text="    Item", font=("Arial", 12, "bold"), 
            fg="black", bg="white", anchor="w", width=15).grid(row=0, column=0, padx=5, pady=3)
        Label(bill_frame, text="Qty", font=("Arial", 12, "bold"), 
            fg="black", bg="white", anchor="center", width=5).grid(row=0, column=1, padx=5, pady=3)
        Label(bill_frame, text="Price", font=("Arial", 12, "bold"), 
            fg="black", bg="white", anchor="e", width=10).grid(row=0, column=2, padx=5, pady=3)
        
        # dotted line after the column names are defined
        Label(bill_frame, text="-" * 70, fg="black", bg="white").grid(row=1, column=0, columnspan=3)
        
        total_price = 0
        row_num = 2
        
        # loading all the items from the cart
        for item, (price, qty) in cart.items():
            Label(bill_frame, text=item, font=("Arial", 11), 
                fg="black", bg="white", anchor="w", width=15).grid(row=row_num, column=0, padx=5, pady=2)
            Label(bill_frame, text=str(qty), font=("Arial", 11), 
                fg="black", bg="white", anchor="center", width=5).grid(row=row_num, column=1, padx=5, pady=2)
            Label(bill_frame, text=str(qty * price), font=("Arial", 11), 
                fg="black", bg="white", anchor="e", width=10).grid(row=row_num, column=2, padx=5, pady=2)
            total_price += qty * price
            row_num += 1
        
        # dotted line after the items list 
        Label(bill_frame, text="-" * 70, fg="black", bg="white").grid(row=row_num, column=0, columnspan=3)
        
        # last line for total bill
        Label(bill_frame, text="   Total:", font=("Arial", 12, "bold"), 
            fg="green", bg="white", anchor="w", width=15).grid(row=row_num+1, column=0, padx=5, pady=5)
        Label(bill_frame, text="₹", font=("Arial", 12, "bold"), 
            fg="green", bg="white", anchor="center", width=5).grid(row=row_num+1, column=1, padx=5, pady=5)
        Label(bill_frame, text=str(total_price), font=("Arial", 12, "bold"), 
            fg="green", bg="white", anchor="e", width=10).grid(row=row_num+1, column=2, padx=5, pady=5)
        
        # Proceed to Pay Button
        Button(bill_win, text="Proceed to Pay", font=("Arial", 12), 
            bg="#006B64", fg="white", command=lambda: process_payment(total_price)).pack(pady=10)



    # search function
    def filter_items(query):
        query = query.lower()

        # ✅ Remove only item sections, keeping search bar & back button
        for widget in section_frame.winfo_children():
            if widget not in [search_frame, back_button_frame]:  # Keep search bar & back button
                widget.destroy()

        has_results = False  # Track if any items are found

        row_num = 1  # Start placing results from row 1 (row 0 is for search bar)
        for section, section_items in items.items():
            filtered_items = [item for item in section_items if query in item["name"].lower()]
            if not filtered_items:
                continue  # Skip empty sections

            has_results = True  # Found at least one match

            # ✅ Use grid() instead of pack()
            parent_frame = Frame(section_frame)
            parent_frame.grid(row=row_num, column=0, sticky="ew", pady=10)

            Label(parent_frame, text=section, font=("Arial", 16), 
                fg="white", bg="#006B64").grid(row=0, column=0, columnspan=5, sticky="ew", pady=5)

            count = 0
            row_frame = Frame(parent_frame)
            row_frame.grid(row=1, column=0, columnspan=5, sticky="ew", pady=5)

            for item in filtered_items:
                if count % 5 == 0:  # New row every 5 items
                    row_frame = Frame(parent_frame)
                    row_frame.grid(row=(count // 5) + 1, column=0, columnspan=5, sticky="ew", pady=10)

                frame = Frame(row_frame, borderwidth=2, relief="ridge", padx=15, pady=15, width=200, height=180)
                frame.grid(row=0, column=count % 5, padx=10, pady=10, sticky="nsew")

                img_path = item["image"] if os.path.exists(item["image"]) else "placeholder.png"
                img = Image.open(img_path).resize((80, 80), Image.LANCZOS)
                img = ImageTk.PhotoImage(img)
                img_label = Label(frame, image=img)
                img_label.image = img
                img_label.grid(row=0, column=0, columnspan=3, pady=5)

                Label(frame, text=item['name'], font=("Arial", 12, "bold")).grid(row=1, column=0, columnspan=3)

                qty_var = StringVar(value="0")

                qty_frame = Frame(frame)
                qty_frame.grid(row=2, column=0, columnspan=3, pady=5)

                Label(qty_frame, text=f"₹{item['price']}", font=("Arial", 12, "bold"), fg="green").grid(row=0, column=0, padx=15)

                Button(qty_frame, text="+", command=lambda iv=qty_var: iv.set(int(iv.get()) + 1)).grid(row=0, column=1)
                Label(qty_frame, textvariable=qty_var, font=("Arial", 12)).grid(row=0, column=2, padx=5)
                Button(qty_frame, text="-", command=lambda iv=qty_var: iv.set(max(0, int(iv.get()) - 1))).grid(row=0, column=3)

                Button(qty_frame, text="Add", command=lambda i=item['name'], p=item['price'], iv=qty_var: update_cart(i, p, iv), 
                    bg="#006B64", fg="white").grid(row=0, column=4, padx=10)

                count += 1

            row_num += 1  # Move to the next row for next section

        # ✅ If no items were found, show a message
        if not has_results:
            Label(section_frame, text="No results found", font=("Arial", 14, "bold"), fg="red").grid(row=row_num, column=0, pady=20)

        # ✅ Show Back Button
        back_button_frame.grid(row=row_num + 1, column=0, pady=5)


    # reset the search when back button is pressed
    def reset_search():
        search_var.set("")  # Clear search input
        back_button_frame.grid_remove()  # Hide back button
        filter_items("")  # Show all items again


    # Main Frame
    main_frame = Frame(root)
    main_frame.pack(fill="both", expand=True)
    
    # Create Canvas for scrolling
    canvas = Canvas(main_frame)
    scrollbar = Scrollbar(main_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = Frame(canvas)
    
    # Configure scrolling
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # ✅ Enable Mouse Scroll
    canvas.bind_all("<MouseWheel>", lambda event: canvas.yview_scroll(-1*(event.delta//120), "units"))  # Windows
    canvas.bind_all("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  # Linux Scroll Up
    canvas.bind_all("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))  # Linux Scroll Down
    
    # Frame inside the scrollable canvas
    section_frame = Frame(scrollable_frame)
    section_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    
    # Ensure expansion inside section_frame
    section_frame.columnconfigure(0, weight=1)
    section_frame.rowconfigure(1, weight=1)  # Allow rows below search to expand

    search_var = StringVar()
    
    # ✅ Create a search frame below the heading, spanning in x-direction
    search_frame = Frame(section_frame)
    search_frame.grid(row=0, column=0, sticky="ew", pady=5, padx=10)
    
    # ✅ Configure column weights to make the search entry expand fully
    search_frame.columnconfigure(1, weight=1)  # Expand the entry field
    
    # ✅ Search Label
    Label(search_frame, text="Search:", font=("Helvica", 12)).grid(row=0, column=0, sticky="w", padx=(0, 5))
    
    # ✅ Search Entry (Input field) - Now expands fully
    search_entry = Entry(search_frame, textvariable=search_var, font=("Helvica", 12))
    search_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=10)  # ✅ Expanded input field
    
    # ✅ Search Button
    Button(search_frame, text="Search", command=lambda: filter_items(search_var.get()), 
        bg="#006B64", fg="white").grid(row=0, column=2, sticky="e", padx=(0, 0))

    
    # ✅ Back Button (initially hidden)
    back_button_frame = Frame(section_frame)
    Button(back_button_frame, text="Back", command=reset_search, bg="red", fg="white").pack()
    


    
    # placing the item frames
    fast_food_frame = Frame(section_frame)
    fast_food_frame.grid(row=1, column=0, sticky="ew")
    snacks_frame = Frame(section_frame)
    snacks_frame.grid(row=2, column=0, sticky="ew")
    beverages_frame = Frame(section_frame)
    beverages_frame.grid(row=3, column=0, sticky="ew")


    # adding the items in respected frames
    items = {
        "Fast Food": [
            {"name": "Onion Samosa(3)", "price": 20, "image": "osamosa.png"},
            {"name": "aloo Samosa", "price": 15, "image": "aloo.png"},
            {"name": "Veg Puff", "price": 25, "image": "veg.png"},
            {"name": "Egg Puff", "price": 35, "image": "egg.png"},
            {"name": "Chicken Puff", "price": 45, "image": "chicken.png"},
            {"name": "Veg Noodles", "price": 35, "image": "veg_nd.png"},
            {"name": "Egg Noodles", "price": 45, "image": "egg_nd.png"},
            {"name": "Chicken Noodles", "price": 65, "image": "chicken_nd.png"},
            {"name": "Veg Shawarma", "price": 125, "image": "veg_sh.png"},
            {"name": "Paneer Shawarma", "price": 135, "image": "paneer.png"},
            {"name": "Chicken Shawarma", "price": 145, "image": "chicken_sh.png"},
            {"name": "Chicken Shawarma + Chips", "price": 155, "image": "chips.png"},
            {"name": "Extra Chicken Shawarma", "price": 165, "image": "ex_sh.png"}
            ],
        "Snacks": [
            {"name": "Lays Classic", "price": 20, "image": "lays-classic.png"},
            {"name": "Kurkure", "price": 20, "image": "kurkure.png"},
            {"name": "Nachos", "price": 40, "image": "nachos.png"},
            {"name": "5 Star", "price": 20, "image": "5star.png"},
            {"name": "Diary Milk Silk", "price": 100, "image": "silk.png"},
            {"name": "Snickers", "price": 40, "image": "snickers.png"},
            {"name": "Oreo", "price": 30, "image": "oreo.png"},
            {"name": "KitKat", "price": 25, "image": "kitkat.png"},
            {"name": "Popcorn", "price": 35, "image": "popcorn.png"},],
        "Beverages": [
            {"name": "Tea", "price": 10, "image": "tea.png"},
            {"name": "Ginger Tea", "price": 15, "image": "ginger.png"},
            {"name": "Coffee", "price": 10, "image": "coffee.png"},
            {"name": "Coke", "price": 35, "image": "coke.png"},
            {"name": "Zero Coke", "price": 40, "image": "zerocoke.png"},
            {"name": "Pepsi", "price": 35, "image": "pepsi.png"},
            {"name": "Pepsi Zero Sugar", "price": 45, "image": "zeropepsi.png"},
            {"name": "Sprite", "price": 30, "image": "sprite.png"},
            {"name": "Fanta", "price": 35, "image": "fanta.png"},
            {"name": "Thumsup", "price": 40, "image": "thumsup.png"},
            {"name": "Mountaindew", "price": 40, "image": "mountain.png"},
            {"name": "Fruit Juice", "price": 45, "image": "fruit.png"},]
    }


    # Displaying all the items with the common GUI elements
    for section, section_items in items.items():
        # Determine which frame to use
        if section == "Snacks":
            parent_frame = snacks_frame
        elif section == "Beverages":
            parent_frame = beverages_frame
        elif section == "Fast Food":
            parent_frame = fast_food_frame
        else:
            continue  # Skip unknown sections
    
        Label(parent_frame, text=section, font=("Arial", 16), fg="white", bg="#006B64").pack(fill='x', pady=5)
    
        # Adding rows of items
        row_frame = Frame(parent_frame)
        row_frame.pack(fill='x', pady=5)
        count = 0
    
        for item in section_items:
            if count % 5 == 0:
                row_frame = Frame(parent_frame)
                row_frame.pack(fill='x', pady=10)
    
            frame = Frame(row_frame, borderwidth=2, relief="ridge", padx=15, pady=15, width=200, height=180)
            frame.pack(side='left', padx=10, pady=10, fill='both', expand=True)
    
            # Loading the dedicated image for the item in the list
            img_path = item["image"] if os.path.exists(item["image"]) else "placeholder.png"
            img = Image.open(img_path).resize((80, 80), Image.LANCZOS)
            img = ImageTk.PhotoImage(img)
            img_label = Label(frame, image=img)
            img_label.image = img
            img_label.pack(pady=5)
    
            # Labeling the image
            Label(frame, text=item['name'], font=("Arial", 12, "bold")).pack()
    
            qty_var = StringVar(value="0")
    
            # Creating a quantity frame to add the price and quantity controls
            qty_frame = Frame(frame)
            qty_frame.pack(side="bottom", pady=5)
    
            # Labeling the price for that item
            Label(qty_frame, text=f"₹{item['price']}", font=("Arial", 12, "bold"), fg="green").pack(side='left', padx=15)
    
            Button(qty_frame, text="+", command=lambda iv=qty_var: iv.set(int(iv.get()) + 1)).pack(side='left')
            Label(qty_frame, textvariable=qty_var, font=("Arial", 12)).pack(side='left', padx=5)
            Button(qty_frame, text="-", command=lambda iv=qty_var: iv.set(max(0, int(iv.get()) - 1))).pack(side='left')
    
            # Add button to finally add the item to the cart with the quantity selected
            Button(qty_frame, text="Add", command=lambda i=item['name'], 
                p=item['price'], iv=qty_var: update_cart(i, p, iv), bg="#006B64", fg="white").pack(side='left', padx=10)
            
            count += 1
    
    # Button to proceed to bill summary and payment 
    Button(root, text="View Cart", command=show_bill, bg="#006B64", fg="white").pack(pady=10)






    def check_and_deduct_balance(hex_value, total_amount, excel_filename="users.xlsx"):
        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(excel_filename)
            sheet = workbook.active
    
            # Get column indexes dynamically
            header = {cell.value.strip().lower(): idx for idx, cell in enumerate(sheet[1])}
            required_columns = ["rfid", "balance", "status"]
            column_indexes = {}
    
            # Ensure required columns exist
            for col_name in required_columns:
                if col_name in header:
                    column_indexes[col_name] = header[col_name]
                else:
                    print(f"❌ Error: Column '{col_name}' not found in the sheet.")
                    return False
    
            # Ensure "Total Amount Spent" column exists
            total_spent_col_name = "total amount spent"
            if total_spent_col_name in header:
                total_spent_col = header[total_spent_col_name]
            else:
                # Add a new column for "Total Amount Spent" if not found
                total_spent_col = len(header)
                sheet.cell(row=1, column=total_spent_col + 1, value=total_spent_col_name.capitalize())
    
            # Get column indexes
            rfid_col = column_indexes["rfid"]
            balance_col = column_indexes["balance"]
            status_col = column_indexes["status"]
    
            user_found = False
            balance_sufficient = False
    
            for row in sheet.iter_rows(min_row=2):  # Skip header row
                rfid_cell = row[rfid_col]
                balance_cell = row[balance_col]
                status_cell = row[status_col]
    
                if str(rfid_cell.value).strip().lower() == hex_value.strip().lower():
                    user_found = True
                    status = str(status_cell.value).strip().lower()
    
                    if status != "true":  # Ensure RFID payment is enabled
                        show_custom_message("Payment Failed", "⚠️ RFID Payment is Disabled for this card!", "#dc3545")
                        return False
    
                    balance = float(balance_cell.value)
    
                    if balance >= total_amount:
                        balance_cell.value = balance - total_amount  # Deduct balance
                        print("✅ Payment successful! New Balance:", balance_cell.value)
                        balance_sufficient = True
    
                        # ✅ Update "Total Amount Spent" ONLY IF balance deduction was successful
                        total_spent_cell = row[total_spent_col]
                        if total_spent_cell.value is None or total_spent_cell.value == "":
                            total_spent_cell.value = total_amount  # First transaction
                        else:
                            total_spent_cell.value = float(total_spent_cell.value) + total_amount  # Add to previous amount
    
                    else:
                        print("❌ Insufficient balance!")
                        show_custom_message("Payment Failed", "⚠️ Insufficient Balance!", "#dc3545")
                        return False
                    break
    
            # If user not found, show message and start QR payment
            if not user_found:
                print(f"❌ User not found! Please register.\nYour RFID Value: {hex_value}")
    
                def start_qr_after_message():
                    time.sleep(2)
                    print("⚠️ Switching to QR code payment since RFID failed.")
                    generate_qr(total_amount)
                    show_qr_code()
                    subprocess.Popen([sys.executable, "login.py"])  # ✅ Ensures login.py runs with the same Python interpreter
    
                threading.Thread(target=lambda: show_custom_message("User Not Found", "⚠️ The RFID user was not found. Please register.", "#dc3545")).start()
                threading.Thread(target=start_qr_after_message).start()
                return False  
    
            # If balance was deducted successfully, save the updated Excel sheet
            if user_found and balance_sufficient:
                workbook.save(excel_filename)
                print("✅ Excel updated successfully!")
                return True
    
        except Exception as e:
            print(f"⚠️ Error handling Excel file: {e}")
            return False
    
    
    



    # Custom message after the payment was done(either successful or not)
    def show_custom_message(title, message, color, qr_price=None):
        root = tk.Tk()
        root.withdraw()  # Hide the main window
    
        popup = tk.Toplevel()
        popup.title(title)
        popup.configure(bg=color)
    
        # Set popup size and center it
        width, height = 600, 150
        screen_width = popup.winfo_screenwidth()
        screen_height = popup.winfo_screenheight()
        x_pos = (screen_width - width) // 2
        y_pos = (screen_height - height) // 2
        popup.geometry(f"{width}x{height}+{x_pos}+{y_pos}")
    
        # Message label
        label = tk.Label(popup, text=message, fg="white", bg=color, font=("Arial", 14, "bold"), padx=20, pady=20)
        label.pack()
    
        # OK button to close the window
        button = tk.Button(popup, text="OK", command=popup.destroy, font=("Arial", 12, "bold"), bg="white", fg="black")
        button.pack(pady=10)
    
        # ✅ If QR payment is needed, schedule it after 1.25 seconds
        if qr_price is not None:
            popup.after(1250, lambda: (generate_qr(qr_price), show_qr_code()))
    
        popup.mainloop()



    # generating the qr and storing it as an image
    # and midifying evertime the total price changes
    def generate_qr(total_price):
        if total_price <= 0:  # Ensure the price is valid
            print("Invalid total price for QR generation.")
            return
    
        qr_data = total_price
        qr = qrcode.make(qr_data)
        qr.save("payment_qr.png")  # Save the QR code
        #qr.show()  # Open the QR code image
        print("✅ QR Code Generated Successfully!")


    # function to show the obtained qr code when needed
    def show_qr_code():
        qr_window = tk.Toplevel()
        qr_window.title("QR Code")
        qr_window.geometry("300x300")  # Adjust window size

        # adding the logo at the top of the window
        logo_img = Image.open("logo.png")
        logo_img = ImageTk.PhotoImage(logo_img)
        qr_window.iconphoto(False, logo_img)
    
        try:
            qr_img = Image.open("payment_qr.png")
            qr_img = qr_img.resize((250, 250), Image.LANCZOS)  # Resize to fit
            qr_img = ImageTk.PhotoImage(qr_img)
    
            qr_label = Label(qr_window, image=qr_img)
            qr_label.image = qr_img  # Keep a reference
            qr_label.pack(pady=10)
    
        except Exception as e:
            Label(qr_window, text="QR Code not found!", fg="red").pack()
            print("Error displaying QR:", e)
    
        qr_window.mainloop()


    # payment process function using RFID or QR
    def process_payment(total_price):
        hex_value = rfid()

        if hex_value == "MODULE_NOT_FOUND":  # ✅ RFID module not detected
            print("❌ No RFID Module Found. Proceeding with QR code payment.")
            
            # ✅ Show message & trigger QR code after 2 seconds
            show_custom_message("RFID Module Not Found", "⚠️ RFID Module is not connected!", "#dc3545", qr_price=total_price)
            return    

        if hex_value is None:  # ✅ No card placed, but module exists
            print("⚠️ No RFID card detected.")
            show_custom_message("Payment Failed", "No RFID Card Detected", "#dc3545")
            return  

        # ✅ Attempt RFID payment
        result = check_and_deduct_balance(hex_value, total_price)

        if result:
            show_custom_message("Payment Successful", "Your payment was successful!", "#006B64")  # Green

    
    root.mainloop()

gitam_mart()
